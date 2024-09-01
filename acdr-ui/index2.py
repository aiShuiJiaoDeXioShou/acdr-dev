from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 加载已生成的 Excel 文件
file_path = './功能模块_功能项_描述_优先级_报价.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# 设置标题行的字体、填充颜色和对齐方式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:  # 第一行通常是标题行
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

# 设置内容行的字体和边框
content_font = Font(color="000000")
for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
    for cell in row:
        cell.font = content_font
        cell.alignment = alignment
        cell.border = thin_border

# 自动调整列宽
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列的字母
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# 保存美化后的 Excel 文件
wb.save(file_path)
